#!/usr/bin/env python3
"""
Excel (.xlsx) -> CSV / JSON exporter for Unity & design pipelines.

Convention (recommended):
  - Row 1: field keys used in code (e.g. id, name, hp) — English, no spaces.
  - Row 2: optional comment row for humans — skipped in export if --skip-row2.
  - Row 3+: data rows.

Empty rows are skipped. UTF-8 with BOM for CSV (Excel-friendly on Windows).
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("Missing dependency: openpyxl", file=sys.stderr)
    print("  pip install -r requirements.txt", file=sys.stderr)
    sys.exit(1)


def slug_key(s: str) -> str:
    """Make a safe JSON key from header cell text."""
    if s is None:
        return ""
    t = str(s).strip()
    if not t:
        return ""
    # Allow common id/name patterns; replace spaces with underscore
    t = re.sub(r"\s+", "_", t)
    t = re.sub(r"[^\w\u4e00-\u9fff]", "_", t)
    t = re.sub(r"_+", "_", t).strip("_")
    return t or "col"


def row_is_empty(values: list) -> bool:
    return all(v is None or (isinstance(v, str) and not v.strip()) for v in values)


def cell_to_json_value(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    # openpyxl may give datetime
    if hasattr(v, "isoformat"):
        try:
            return v.isoformat()
        except Exception:
            pass
    s = str(v).strip()
    if s == "":
        return None
    # Try int / float for game data convenience
    if re.fullmatch(r"-?\d+", s):
        try:
            return int(s)
        except ValueError:
            pass
    if re.fullmatch(r"-?\d+\.\d+", s):
        try:
            return float(s)
        except ValueError:
            pass
    low = s.lower()
    if low in ("true", "false"):
        return low == "true"
    return s


def export_sheet_to_rows(ws, skip_row2: bool):
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return [], []

    if skip_row2:
        try:
            next(rows_iter)  # discard comment / type row
        except StopIteration:
            pass

    keys_raw = [slug_key(h) for h in header]
    # Ensure unique keys
    count: dict[str, int] = {}
    keys = []
    for k in keys_raw:
        base = k or "col"
        c = count.get(base, 0)
        final = base if c == 0 else f"{base}_{c}"
        count[base] = c + 1
        keys.append(final)

    data = []
    for row in rows_iter:
        if row_is_empty(row):
            continue
        obj = {}
        for i, key in enumerate(keys):
            if i >= len(row):
                obj[key] = None
            else:
                obj[key] = cell_to_json_value(row[i])
        data.append(obj)
    return keys, data


def write_csv(path: Path, keys: list[str], rows: list[dict], utf8_bom: bool):
    encoding = "utf-8-sig" if utf8_bom else "utf-8"
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding=encoding) as f:
        w = csv.DictWriter(f, fieldnames=keys, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            # CSV: stringify for consistency
            out = {}
            for k in keys:
                v = r.get(k)
                if v is None:
                    out[k] = ""
                else:
                    out[k] = v if isinstance(v, str) else json.dumps(v, ensure_ascii=False)
            w.writerow(out)


def main():
    p = argparse.ArgumentParser(description="Export .xlsx to CSV and/or JSON.")
    p.add_argument("input", type=Path, help="Path to .xlsx file")
    p.add_argument(
        "-o",
        "--out-dir",
        type=Path,
        default=None,
        help="Output directory (default: same folder as xlsx, subfolder 'export')",
    )
    p.add_argument(
        "--format",
        choices=("csv", "json", "both"),
        default="both",
        help="Output format",
    )
    p.add_argument(
        "--sheets",
        nargs="*",
        default=None,
        help="Sheet names to export (default: all)",
    )
    p.add_argument(
        "--skip-row2",
        action="store_true",
        help="Treat row 2 as comments / types and skip it",
    )
    p.add_argument(
        "--unity-json",
        action="store_true",
        help="JSON shape: one file per sheet with {\"list\":[...]} for JsonUtility",
    )
    p.add_argument(
        "--no-utf8-bom-csv",
        action="store_true",
        help="Write CSV as UTF-8 without BOM",
    )
    p.add_argument(
        "--json-basename",
        default=None,
        metavar="NAME",
        help="JSON filename without .json (only when exporting exactly one sheet)",
    )
    args = p.parse_args()

    xlsx = args.input.resolve()
    if not xlsx.is_file():
        print(f"File not found: {xlsx}", file=sys.stderr)
        sys.exit(1)

    out_dir = args.out_dir
    if out_dir is None:
        out_dir = xlsx.parent / "export"
    out_dir = out_dir.resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx, data_only=True, read_only=True)
    names = args.sheets if args.sheets else wb.sheetnames
    if args.json_basename and len(names) != 1:
        print(
            "Error: --json-basename requires exactly one sheet (use --sheets MySheet).",
            file=sys.stderr,
        )
        wb.close()
        sys.exit(1)

    json_override = args.json_basename
    if json_override:
        json_override = re.sub(r"[^\w\u4e00-\u9fff._-]+", "_", json_override).strip(
            "._-"
        ) or "table"

    for name in names:
        if name not in wb.sheetnames:
            print(f"Warning: sheet not found, skip: {name}", file=sys.stderr)
            continue
        ws = wb[name]
        keys, rows = export_sheet_to_rows(ws, skip_row2=args.skip_row2)
        safe_name = re.sub(r"[^\w\u4e00-\u9fff]+", "_", name).strip("_") or "sheet"

        if args.format in ("csv", "both"):
            csv_path = out_dir / f"{safe_name}.csv"
            write_csv(csv_path, keys, rows, utf8_bom=not args.no_utf8_bom_csv)
            print(f"CSV: {csv_path}")

        if args.format in ("json", "both"):
            json_stem = json_override if json_override else safe_name
            json_path = out_dir / f"{json_stem}.json"
            payload = rows
            if args.unity_json:
                payload = {"list": rows}
            json_path.write_text(
                json.dumps(payload, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
            print(f"JSON: {json_path}")

    wb.close()
    print("Done.")


if __name__ == "__main__":
    main()
